#!/usr/bin/env python3
"""
Costruisce Excel 3-fogli con catalogo completo di cosa NEXO può monitorare
nella Suite ACG, numeri attuali (letti da Firestore) e breakdown per tecnico.

Output: scripts/report_monitoraggio.xlsx
"""
from __future__ import annotations

import os
import sys
import datetime as dt
from collections import defaultdict

import firebase_admin
from firebase_admin import credentials, firestore
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── Firebase apps ─────────────────────────────────────────────
def _init_app(project_id: str, name: str):
    try:
        return firebase_admin.get_app(name)
    except ValueError:
        return firebase_admin.initialize_app(credentials.ApplicationDefault(), {"projectId": project_id}, name=name)

app_cosmina = _init_app("garbymobile-f89ac", "cosmina")
app_gzt     = _init_app("guazzotti-tec",     "guazzotti")
app_nexo    = _init_app("nexo-hub-15f2d",    "nexo")
db_cosmina  = firestore.client(app_cosmina)
db_gzt      = firestore.client(app_gzt)
db_nexo     = firestore.client(app_nexo)

# ─── Catalogo monitoraggio ─────────────────────────────────────
# Ogni entry ha: categoria, elemento, descrizione, fonte, collection,
# frequenza_suggerita, severita, badge_chronos, query_fn (o None)
def _count(db, coll, filter_fn=None, limit=None):
    """Conta docs. Per filter_fn: se None usa count().get() aggregation.
    Se c'è filter_fn, fetcha (con cap 3000) e filtra in-memory."""
    try:
        q = db.collection(coll)
        if filter_fn is None:
            return q.count().get()[0][0].value
        limit = limit or 3000
        docs = q.limit(limit).stream()
        return sum(1 for d in docs if filter_fn(d.to_dict() or {}))
    except Exception as e:
        return f"ERR: {str(e)[:50]}"

def _sum_field(db, coll, field, filter_fn=None, limit=3000):
    try:
        total = 0.0
        for d in db.collection(coll).limit(limit).stream():
            v = d.to_dict() or {}
            if filter_fn and not filter_fn(v):
                continue
            x = v.get(field)
            if isinstance(x, (int, float)):
                total += x
        return total
    except Exception as e:
        return f"ERR: {str(e)[:50]}"

def _now():
    return dt.datetime.now(dt.timezone.utc)

def _days_ago(n):
    return _now() - dt.timedelta(days=n)

def _parse_date(v):
    if v is None: return None
    # Firestore Timestamp / datetime
    if isinstance(v, dt.datetime):
        return v if v.tzinfo else v.replace(tzinfo=dt.timezone.utc)
    # DatetimeWithNanoseconds (Google)
    if hasattr(v, "to_pydatetime"):
        try:
            d = v.to_pydatetime()
            return d if d.tzinfo else d.replace(tzinfo=dt.timezone.utc)
        except Exception:
            pass
    if isinstance(v, str):
        for fmt in ("%Y-%m-%dT%H:%M:%S", "%Y-%m-%d", "%d/%m/%Y"):
            try:
                s = v[:19] if fmt.startswith("%Y-%m-%dT") else v[:10]
                d = dt.datetime.strptime(s, fmt)
                return d.replace(tzinfo=dt.timezone.utc)
            except Exception:
                continue
    return None

# ─── Catalogo (riga × metrica) ─────────────────────────────────
# "query" è una callable (db_cosmina, db_gzt, db_nexo) -> numero | None (placeholder)
CATALOG = [
    # ─── CAMPAGNE ───────────────────────────────────────────────
    dict(cat="Campagne", el="Campagne attive totali", desc="Numero campagne non archiviate in cosmina_campagne",
         fonte="COSMINA", coll="cosmina_campagne", freq="giornaliera", sev="info", badge=True,
         warn=None, crit=None,
         query=lambda: _count(db_cosmina, "cosmina_campagne",
                              lambda v: v.get("stato") != "archiviata" and not v.get("archiviata"))),
    dict(cat="Campagne", el="SPEGNIMENTO 2026 — totale", desc="Interventi totali campagna Spegnimento",
         fonte="COSMINA", coll="bacheca_cards", freq="giornaliera", sev="info", badge=True,
         warn=None, crit=None,
         query=lambda: _count(db_cosmina, "bacheca_cards",
                              lambda v: v.get("campagna_nome") == "SPEGNIMENTO 2026")),
    dict(cat="Campagne", el="SPEGNIMENTO 2026 — scaduti", desc="Interventi Spegnimento con stato=aperto e due<oggi",
         fonte="COSMINA", coll="bacheca_cards", freq="giornaliera", sev="warning", badge=True,
         warn=5, crit=20,
         query=lambda: _count(db_cosmina, "bacheca_cards", lambda v: (
             v.get("campagna_nome") == "SPEGNIMENTO 2026"
             and str(v.get("stato")).lower() == "aperto"
             and _parse_date(v.get("due")) and _parse_date(v.get("due")) < _now()
         ))),
    dict(cat="Campagne", el="Letture WalkBy ACG FS 2026 — totale", desc="Totale interventi campagna WalkBy ACG",
         fonte="COSMINA", coll="bacheca_cards", freq="giornaliera", sev="info", badge=True,
         warn=None, crit=None,
         query=lambda: _count(db_cosmina, "bacheca_cards",
                              lambda v: v.get("campagna_nome") == "Letture WalkBy ACG FS 2026")),
    dict(cat="Campagne", el="Letture WalkBy GZT FS 2026 — totale", desc="Totale interventi campagna WalkBy Guazzotti",
         fonte="COSMINA", coll="bacheca_cards", freq="giornaliera", sev="info", badge=True,
         warn=None, crit=None,
         query=lambda: _count(db_cosmina, "bacheca_cards",
                              lambda v: v.get("campagna_nome") == "Letture WalkBy GZT FS 2026")),
    dict(cat="Campagne", el="Campagne % completamento medio", desc="% media completati su totali per campagne attive",
         fonte="COSMINA", coll="bacheca_cards", freq="giornaliera", sev="info", badge=False,
         warn=None, crit=None, query=None),
    dict(cat="Campagne", el="Report campagne generati", desc="Docs in cosmina_campagne_reports",
         fonte="COSMINA", coll="cosmina_campagne_reports", freq="settimanale", sev="info", badge=False,
         warn=None, crit=None, query=lambda: _count(db_cosmina, "cosmina_campagne_reports")),

    # ─── INTERVENTI (bacheca_cards) ─────────────────────────────
    dict(cat="Interventi", el="Interventi attivi (bacheca INTERVENTI)", desc="Cards listName=INTERVENTI, inBacheca=true",
         fonte="COSMINA", coll="bacheca_cards", freq="continua", sev="info", badge=True,
         warn=None, crit=None,
         query=lambda: _count(db_cosmina, "bacheca_cards", lambda v: (
             v.get("listName") == "INTERVENTI" and v.get("inBacheca") is not False
         ))),
    dict(cat="Interventi", el="Interventi aperti totali", desc="Cards stato!=chiuso, in qualsiasi listName",
         fonte="COSMINA", coll="bacheca_cards", freq="continua", sev="info", badge=True,
         warn=None, crit=None,
         query=lambda: _count(db_cosmina, "bacheca_cards",
                              lambda v: str(v.get("stato", "")).lower() not in ("chiuso", "completato"))),
    dict(cat="Interventi", el="Interventi oggi", desc="Cards con due oggi, stato aperto",
         fonte="COSMINA", coll="bacheca_cards", freq="continua", sev="info", badge=True,
         warn=None, crit=None,
         query=lambda: _count(db_cosmina, "bacheca_cards", lambda v: (
             v.get("listName") == "INTERVENTI"
             and v.get("inBacheca") is not False
             and _parse_date(v.get("due")) and
             _parse_date(v.get("due")).date() == dt.datetime.utcnow().date()
         ))),
    dict(cat="Interventi", el="Interventi scaduti", desc="Cards INTERVENTI con due<oggi, stato aperto",
         fonte="COSMINA", coll="bacheca_cards", freq="continua", sev="critical", badge=True,
         warn=30, crit=100,
         query=lambda: _count(db_cosmina, "bacheca_cards", lambda v: (
             v.get("listName") == "INTERVENTI"
             and v.get("inBacheca") is not False
             and str(v.get("stato", "")).lower() == "aperto"
             and _parse_date(v.get("due")) and _parse_date(v.get("due")) < _now() - dt.timedelta(days=1)
         ))),
    dict(cat="Interventi", el="Interventi senza tecnico", desc="Cards senza techName/techNames assegnati",
         fonte="COSMINA", coll="bacheca_cards", freq="giornaliera", sev="warning", badge=True,
         warn=10, crit=40,
         query=lambda: _count(db_cosmina, "bacheca_cards", lambda v: (
             v.get("listName") == "INTERVENTI" and v.get("inBacheca") is not False
             and not v.get("techName") and not (v.get("techNames") or [])
         ))),
    dict(cat="Interventi", el="Interventi senza data (da programmare)", desc="Cards INTERVENTI con due=None",
         fonte="COSMINA", coll="bacheca_cards", freq="giornaliera", sev="warning", badge=True,
         warn=20, crit=60,
         query=lambda: _count(db_cosmina, "bacheca_cards", lambda v: (
             v.get("listName") == "INTERVENTI" and v.get("inBacheca") is not False
             and str(v.get("stato", "")).lower() == "aperto"
             and not v.get("due")
         ))),
    dict(cat="Interventi", el="Interventi settimana prossima", desc="Due entro 7 giorni",
         fonte="COSMINA", coll="bacheca_cards", freq="giornaliera", sev="info", badge=True,
         warn=None, crit=None,
         query=lambda: _count(db_cosmina, "bacheca_cards", lambda v: (
             v.get("listName") == "INTERVENTI" and v.get("inBacheca") is not False
             and _parse_date(v.get("due")) and
             _now() <= _parse_date(v.get("due")) <= _now() + dt.timedelta(days=7)
         ))),

    # ─── RTI GUAZZOTTI ──────────────────────────────────────────
    dict(cat="RTI/RTIDF", el="RTI totali", desc="Tutti i rapporti tecnici",
         fonte="Guazzotti", coll="rti", freq="giornaliera", sev="info", badge=False,
         warn=None, crit=None, query=lambda: _count(db_gzt, "rti")),
    dict(cat="RTI/RTIDF", el="RTI in bozza", desc="Rapporti aperti, non chiusi dal tecnico",
         fonte="Guazzotti", coll="rti", freq="giornaliera", sev="warning", badge=True,
         warn=20, crit=60,
         query=lambda: _count(db_gzt, "rti", lambda v: v.get("stato") == "bozza")),
    dict(cat="RTI/RTIDF", el="RTI bozza >7 giorni", desc="Rapporti chiusi dal tecnico da oltre 7g ma ancora bozza",
         fonte="Guazzotti", coll="rti", freq="giornaliera", sev="warning", badge=True,
         warn=15, crit=40,
         query=lambda: _count(db_gzt, "rti", lambda v: (
             v.get("stato") == "bozza"
             and _parse_date(v.get("_lastModified") or v.get("timestamp"))
             and _parse_date(v.get("_lastModified") or v.get("timestamp")) < _days_ago(7)
         ))),
    dict(cat="RTI/RTIDF", el="RTI bozza >90 giorni", desc="Bozze dimenticate da oltre 3 mesi",
         fonte="Guazzotti", coll="rti", freq="settimanale", sev="critical", badge=True,
         warn=5, crit=15,
         query=lambda: _count(db_gzt, "rti", lambda v: (
             v.get("stato") == "bozza"
             and _parse_date(v.get("_lastModified") or v.get("timestamp"))
             and _parse_date(v.get("_lastModified") or v.get("timestamp")) < _days_ago(90)
         ))),
    dict(cat="RTI/RTIDF", el="RTI definito senza RTIDF", desc="RTI pronti per duplicazione in RTIDF — blocco fatturazione",
         fonte="Guazzotti", coll="rti", freq="giornaliera", sev="critical", badge=True,
         warn=50, crit=150,
         query=lambda: _count(db_gzt, "rti", lambda v: v.get("stato") == "definito")),
    dict(cat="RTI/RTIDF", el="RTI con stato 'da_verificare'", desc="Stato anomalo emerso, da chiarire",
         fonte="Guazzotti", coll="rti", freq="settimanale", sev="warning", badge=False,
         warn=5, crit=20,
         query=lambda: _count(db_gzt, "rti", lambda v: v.get("stato") == "da_verificare")),
    dict(cat="RTI/RTIDF", el="RTI senza materiali compilati", desc="Array materiali vuoto (compliance RTI)",
         fonte="Guazzotti", coll="rti", freq="settimanale", sev="info", badge=False,
         warn=None, crit=None,
         query=lambda: _count(db_gzt, "rti", lambda v: not (v.get("materiali") or []))),

    # ─── RTIDF GUAZZOTTI ────────────────────────────────────────
    dict(cat="RTI/RTIDF", el="RTIDF totali", desc="RTI Definitivi generati",
         fonte="Guazzotti", coll="rtidf", freq="giornaliera", sev="info", badge=False,
         warn=None, crit=None, query=lambda: _count(db_gzt, "rtidf")),
    dict(cat="RTI/RTIDF", el="RTIDF inviati (pronti fatturazione)", desc="RTIDF stato=inviato — ricavi bloccati",
         fonte="Guazzotti", coll="rtidf", freq="giornaliera", sev="critical", badge=True,
         warn=10, crit=30,
         query=lambda: _count(db_gzt, "rtidf", lambda v: v.get("stato") == "inviato")),
    dict(cat="RTI/RTIDF", el="RTIDF senza costo_intervento", desc="RTIDF da completare prima di fatturare",
         fonte="Guazzotti", coll="rtidf", freq="giornaliera", sev="critical", badge=True,
         warn=30, crit=100,
         query=lambda: _count(db_gzt, "rtidf", lambda v: (
             (v.get("costo_intervento") is None or v.get("costo_intervento") == 0)
             and v.get("stato") not in ("bozza",)
         ))),
    dict(cat="RTI/RTIDF", el="RTIDF fatturati", desc="RTIDF inclusi in una commessa — completati",
         fonte="Guazzotti", coll="rtidf", freq="settimanale", sev="info", badge=False,
         warn=None, crit=None,
         query=lambda: _count(db_gzt, "rtidf", lambda v: v.get("stato") == "fatturato")),
    dict(cat="RTI/RTIDF", el="RTIDF orfani (senza RTI origine)", desc="RTIDF senza rti_origine_id",
         fonte="Guazzotti", coll="rtidf", freq="settimanale", sev="warning", badge=False,
         warn=1, crit=5,
         query=lambda: _count(db_gzt, "rtidf", lambda v: not v.get("rti_origine_id"))),

    # ─── TICKETS GUAZZOTTI ──────────────────────────────────────
    dict(cat="Interventi", el="Tickets aperti", desc="Tickets stato=aperto | pianificato | in_attesa | da_chiudere",
         fonte="Guazzotti", coll="tickets", freq="continua", sev="info", badge=True,
         warn=None, crit=None,
         query=lambda: _count(db_gzt, "tickets",
                              lambda v: v.get("stato") in ("aperto", "pianificato", "in_attesa", "da_chiudere"))),
    dict(cat="Interventi", el="Tickets aperti >30g", desc="Ticket aperti da più di 30 giorni",
         fonte="Guazzotti", coll="tickets", freq="giornaliera", sev="warning", badge=True,
         warn=30, crit=80,
         query=lambda: _count(db_gzt, "tickets", lambda v: (
             v.get("stato") in ("aperto", "pianificato", "in_attesa", "da_chiudere")
             and _parse_date(v.get("data_apertura") or v.get("timestamp"))
             and _parse_date(v.get("data_apertura") or v.get("timestamp")) < _days_ago(30)
         ))),
    dict(cat="Interventi", el="Tickets aperti senza RTI", desc="Tickets aperti senza rti_inviato",
         fonte="Guazzotti", coll="tickets", freq="giornaliera", sev="warning", badge=True,
         warn=50, crit=120,
         query=lambda: _count(db_gzt, "tickets", lambda v: (
             v.get("stato") in ("aperto", "pianificato", "in_attesa", "da_chiudere")
             and not v.get("rti_inviato") and not v.get("rtiChiusura")
         ))),
    dict(cat="Interventi", el="Tickets chiusi senza RTI", desc="Tickets chiusi senza RTI collegato — non fatturabile",
         fonte="Guazzotti", coll="tickets", freq="settimanale", sev="warning", badge=True,
         warn=20, crit=80,
         query=lambda: _count(db_gzt, "tickets", lambda v: (
             str(v.get("stato", "")).startswith("chiuso")
             and not v.get("rti_inviato") and not v.get("rtiChiusura")
         ))),

    # ─── FATTURAZIONE ───────────────────────────────────────────
    dict(cat="Fatturazione", el="Fatture Guazzotti parsate", desc="Fatture parsate da email (docfin_fatture_guazzotti)",
         fonte="COSMINA", coll="docfin_fatture_guazzotti", freq="giornaliera", sev="info", badge=False,
         warn=None, crit=None, query=lambda: _count(db_cosmina, "docfin_fatture_guazzotti")),
    dict(cat="Fatturazione", el="Fatture non pagate", desc="docfin_fatture_guazzotti stato != pagata",
         fonte="COSMINA", coll="docfin_fatture_guazzotti", freq="giornaliera", sev="warning", badge=True,
         warn=10, crit=30,
         query=lambda: _count(db_cosmina, "docfin_fatture_guazzotti",
                              lambda v: str(v.get("stato", "")).lower() != "pagata")),
    dict(cat="Fatturazione", el="Commesse aperte Guazzotti", desc="Commesse stato != chiusa",
         fonte="Guazzotti", coll="commesse", freq="settimanale", sev="info", badge=False,
         warn=None, crit=None,
         query=lambda: _count(db_gzt, "commesse", lambda v: str(v.get("stato", "")).lower() != "chiusa")),

    # ─── PAGAMENTI ─────────────────────────────────────────────
    dict(cat="Fatturazione", el="Esposizione totale clienti (€)", desc="TotaleEsposizione somma su pagamenti_clienti",
         fonte="Guazzotti", coll="pagamenti_clienti", freq="giornaliera", sev="info", badge=True,
         warn=500000, crit=1000000,
         query=lambda: _sum_field(db_gzt, "pagamenti_clienti", "TotaleEsposizione")),
    dict(cat="Fatturazione", el="Scaduto totale (€)", desc="TotaleScaduto somma su pagamenti_clienti",
         fonte="Guazzotti", coll="pagamenti_clienti", freq="giornaliera", sev="critical", badge=True,
         warn=50000, crit=150000,
         query=lambda: _sum_field(db_gzt, "pagamenti_clienti", "TotaleScaduto")),
    dict(cat="Fatturazione", el="Clienti con scaduto >90g", desc="Clienti con Scaduto90+180+360+oltre > 0",
         fonte="Guazzotti", coll="pagamenti_clienti", freq="giornaliera", sev="warning", badge=True,
         warn=5, crit=20,
         query=lambda: _count(db_gzt, "pagamenti_clienti", lambda v: (
             (v.get("Scaduto90") or 0) + (v.get("Scaduto180") or 0) +
             (v.get("Scaduto360") or 0) + (v.get("ScadutoOltre360") or 0) > 0
         ))),
    dict(cat="Fatturazione", el="Clienti rischio ALTO", desc="riskLevel in pagamenti_clienti = ALTO",
         fonte="Guazzotti", coll="pagamenti_clienti", freq="giornaliera", sev="critical", badge=True,
         warn=3, crit=10,
         query=lambda: _count(db_gzt, "pagamenti_clienti",
                              lambda v: str(v.get("riskLevel", "")).upper() == "ALTO")),

    # ─── SCADENZE / IMPIANTI ────────────────────────────────────
    dict(cat="Scadenze", el="Impianti totali", desc="Totale cosmina_impianti",
         fonte="COSMINA", coll="cosmina_impianti", freq="mensile", sev="info", badge=False,
         warn=None, crit=None, query=lambda: _count(db_cosmina, "cosmina_impianti")),
    dict(cat="Scadenze", el="Impianti con scadenza manutenzione registrata", desc="cosmina_impianti con data_prossima_manutenzione o data_prossimo_contributo",
         fonte="COSMINA", coll="cosmina_impianti", freq="settimanale", sev="info", badge=True,
         warn=None, crit=None,
         query=lambda: _count(db_cosmina, "cosmina_impianti", lambda v: (
             v.get("data_prossima_manutenzione") or v.get("data_prossimo_contributo")
         ))),
    dict(cat="Scadenze", el="Impianti con scadenza <30g", desc="Scadenza entro 30 giorni",
         fonte="COSMINA", coll="cosmina_impianti", freq="giornaliera", sev="warning", badge=True,
         warn=5, crit=20,
         query=lambda: _count(db_cosmina, "cosmina_impianti", lambda v: (
             _parse_date(v.get("data_prossima_manutenzione") or v.get("data_prossimo_contributo"))
             and _now() <= _parse_date(v.get("data_prossima_manutenzione") or v.get("data_prossimo_contributo")) <= _now() + dt.timedelta(days=30)
         ))),
    dict(cat="Scadenze", el="Impianti scaduti non fatti", desc="Scadenza passata senza chiusura",
         fonte="COSMINA", coll="cosmina_impianti", freq="giornaliera", sev="critical", badge=True,
         warn=5, crit=20,
         query=lambda: _count(db_cosmina, "cosmina_impianti", lambda v: (
             _parse_date(v.get("data_prossima_manutenzione") or v.get("data_prossimo_contributo"))
             and _parse_date(v.get("data_prossima_manutenzione") or v.get("data_prossimo_contributo")) < _now()
         ))),
    dict(cat="Scadenze", el="Impianti senza targa CURIT", desc="Impianti senza campo targa compilato",
         fonte="COSMINA", coll="cosmina_impianti", freq="mensile", sev="warning", badge=True,
         warn=20, crit=100,
         query=lambda: _count(db_cosmina, "cosmina_impianti", lambda v: not v.get("targa"))),
    dict(cat="Scadenze", el="Impianti senza CIT associato", desc="Impianti senza doc in cosmina_impianti_cit",
         fonte="COSMINA", coll="cosmina_impianti", freq="mensile", sev="info", badge=False,
         warn=None, crit=None, query=None),
    dict(cat="Scadenze", el="CIT impianti totali", desc="Totale cosmina_impianti_cit (dichiarazioni CURIT)",
         fonte="COSMINA", coll="cosmina_impianti_cit", freq="mensile", sev="info", badge=False,
         warn=None, crit=None, query=lambda: _count(db_cosmina, "cosmina_impianti_cit")),
    dict(cat="Scadenze", el="Impianti con ritardo manutenzione >30g", desc="giorni_ritardo_manutenzione > 30",
         fonte="COSMINA", coll="cosmina_impianti", freq="settimanale", sev="warning", badge=True,
         warn=3, crit=15,
         query=lambda: _count(db_cosmina, "cosmina_impianti",
                              lambda v: (v.get("giorni_ritardo_manutenzione") or 0) > 30)),

    # ─── CLIENTI / CRM ──────────────────────────────────────────
    dict(cat="Clienti", el="Clienti CRM totali", desc="Totale crm_clienti",
         fonte="COSMINA", coll="crm_clienti", freq="mensile", sev="info", badge=False,
         warn=None, crit=None, query=lambda: _count(db_cosmina, "crm_clienti")),
    dict(cat="Clienti", el="Clienti senza contratto attivo", desc="crm_clienti con ac2_contratto=false",
         fonte="COSMINA", coll="crm_clienti", freq="mensile", sev="info", badge=False,
         warn=None, crit=None,
         query=lambda: _count(db_cosmina, "crm_clienti", lambda v: v.get("ac2_contratto") is False)),
    dict(cat="Clienti", el="Clienti Gruppo Badano", desc="Totale crm_gruppo_badano",
         fonte="COSMINA", coll="crm_gruppo_badano", freq="mensile", sev="info", badge=False,
         warn=None, crit=None, query=lambda: _count(db_cosmina, "crm_gruppo_badano")),

    # ─── MAGAZZINO ─────────────────────────────────────────────
    dict(cat="Magazzino", el="Articoli a catalogo", desc="Totale magazzino",
         fonte="COSMINA", coll="magazzino", freq="settimanale", sev="info", badge=False,
         warn=None, crit=None, query=lambda: _count(db_cosmina, "magazzino")),
    dict(cat="Magazzino", el="Articoli sotto scorta minima", desc="magazzino_giacenze quantita<scorta_minima",
         fonte="COSMINA", coll="magazzino_giacenze", freq="giornaliera", sev="critical", badge=True,
         warn=5, crit=20,
         query=lambda: _count(db_cosmina, "magazzino_giacenze", lambda v: (
             (v.get("scorta_minima") or 0) > 0 and (v.get("quantita") or 0) < (v.get("scorta_minima") or 0)
         ))),
    dict(cat="Magazzino", el="Giacenze totali (items)", desc="Totale righe magazzino_giacenze",
         fonte="COSMINA", coll="magazzino_giacenze", freq="settimanale", sev="info", badge=False,
         warn=None, crit=None, query=lambda: _count(db_cosmina, "magazzino_giacenze")),
    dict(cat="Magazzino", el="Movimenti ultimo giorno", desc="magazzino_movimenti creati negli ultimi 24h",
         fonte="COSMINA", coll="magazzino_movimenti", freq="giornaliera", sev="info", badge=False,
         warn=None, crit=None, query=None),

    # ─── EMAIL / IRIS ──────────────────────────────────────────
    dict(cat="Email", el="Email IRIS processate totali", desc="iris_emails totale",
         fonte="NEXO", coll="iris_emails", freq="continua", sev="info", badge=False,
         warn=None, crit=None, query=lambda: _count(db_nexo, "iris_emails")),
    dict(cat="Email", el="Email IRIS urgenti non risposte", desc="iris_emails categoria=URGENTE, status!=risposta",
         fonte="NEXO", coll="iris_emails", freq="oraria", sev="critical", badge=True,
         warn=2, crit=5,
         query=lambda: _count(db_nexo, "iris_emails", lambda v: (
             str(v.get("categoria", "")).upper() == "URGENTE"
             and not v.get("replied_at")
         ))),
    dict(cat="Email", el="Email senza risposta >48h", desc="iris_emails con data<now-48h non risposte",
         fonte="NEXO", coll="iris_emails", freq="oraria", sev="warning", badge=True,
         warn=10, crit=30,
         query=lambda: _count(db_nexo, "iris_emails", lambda v: (
             not v.get("replied_at")
             and _parse_date(v.get("received_at") or v.get("timestamp"))
             and _parse_date(v.get("received_at") or v.get("timestamp")) < _days_ago(2)
         ))),
    dict(cat="Email", el="Email oggi (volume)", desc="iris_emails ricevute oggi",
         fonte="NEXO", coll="iris_emails", freq="continua", sev="info", badge=True,
         warn=None, crit=None,
         query=lambda: _count(db_nexo, "iris_emails", lambda v: (
             _parse_date(v.get("received_at") or v.get("timestamp"))
             and _parse_date(v.get("received_at") or v.get("timestamp")).date() == dt.datetime.utcnow().date()
         ))),

    # ─── TECNICI ───────────────────────────────────────────────
    dict(cat="Tecnici", el="Tecnici ACG attivi", desc="acg_tecnici con attivo=true",
         fonte="COSMINA", coll="acg_tecnici", freq="mensile", sev="info", badge=False,
         warn=None, crit=None,
         query=lambda: _count(db_cosmina, "acg_tecnici", lambda v: v.get("attivo") is True)),
    dict(cat="Tecnici", el="Distribuzione interventi per tecnico", desc="Breakdown nel foglio 'PER TECNICO'",
         fonte="COSMINA", coll="bacheca_cards", freq="giornaliera", sev="info", badge=False,
         warn=None, crit=None, query=None),

    # ─── COMPLIANCE / GDPR ─────────────────────────────────────
    dict(cat="Compliance", el="Consensi GDPR registrati", desc="gdpr_consents totale",
         fonte="COSMINA", coll="gdpr_consents", freq="mensile", sev="info", badge=False,
         warn=None, crit=None, query=lambda: _count(db_cosmina, "gdpr_consents")),
    dict(cat="Compliance", el="Dichiarazioni CURIT scadute", desc="impianti con data_scadenza_dichiarazione < oggi",
         fonte="COSMINA", coll="cosmina_impianti", freq="mensile", sev="critical", badge=True,
         warn=5, crit=20,
         query=lambda: _count(db_cosmina, "cosmina_impianti", lambda v: (
             _parse_date(v.get("data_scadenza_dichiarazione"))
             and _parse_date(v.get("data_scadenza_dichiarazione")) < _now()
         ))),

    # ─── INFRASTRUTTURA (PHARO) ────────────────────────────────
    dict(cat="Infrastruttura", el="Servizi in allarme PHARO", desc="alarm_events con stato=attivo",
         fonte="COSMINA", coll="alarm_events", freq="continua", sev="critical", badge=True,
         warn=1, crit=3,
         query=lambda: _count(db_cosmina, "alarm_events", lambda v: str(v.get("stato", "")).lower() == "attivo")),
    dict(cat="Infrastruttura", el="Allarmi pending", desc="alarm_pending non processati",
         fonte="COSMINA", coll="alarm_pending", freq="continua", sev="warning", badge=True,
         warn=5, crit=20,
         query=lambda: _count(db_cosmina, "alarm_pending")),
    dict(cat="Infrastruttura", el="Audit log ultimi 24h", desc="audit_log ultimi 24h — volume API calls",
         fonte="COSMINA", coll="audit_log", freq="oraria", sev="info", badge=False,
         warn=None, crit=None, query=None),
    dict(cat="Infrastruttura", el="Costi AI DARWIN", desc="darwin_costs — budget AI speso",
         fonte="COSMINA", coll="darwin_costs", freq="giornaliera", sev="info", badge=True,
         warn=None, crit=None, query=lambda: _count(db_cosmina, "darwin_costs")),
    dict(cat="Infrastruttura", el="AI audit log operazioni", desc="ai_audit_log — conteggio interazioni AI app",
         fonte="COSMINA", coll="ai_audit_log", freq="giornaliera", sev="info", badge=False,
         warn=None, crit=None, query=lambda: _count(db_cosmina, "ai_audit_log")),

    # ─── RIPARTITORI / CANTIERI ────────────────────────────────
    dict(cat="Compliance", el="Ripartitori UNI 10200 attivi", desc="cosmina_ripartitori non terminati",
         fonte="COSMINA", coll="cosmina_ripartitori", freq="mensile", sev="info", badge=False,
         warn=None, crit=None,
         query=lambda: _count(db_cosmina, "cosmina_ripartitori", lambda v: (
             "TERMIN" not in str(v.get("avanzamento", "")).upper()
         ))),
    dict(cat="Compliance", el="Ripartitori in scadenza", desc="Ripartitori con tipo_rdo=DEFINITIVO da chiudere",
         fonte="COSMINA", coll="cosmina_ripartitori", freq="mensile", sev="warning", badge=False,
         warn=None, crit=None,
         query=lambda: _count(db_cosmina, "cosmina_ripartitori", lambda v: (
             v.get("tipo_rdo") == "DEFINITIVO"
             and str(v.get("stato_commerciale", "")).upper() != "PAGATO CHIUSO"
         ))),
    dict(cat="Compliance", el="Cantieri attivi", desc="cantieri totale",
         fonte="COSMINA", coll="cantieri", freq="mensile", sev="info", badge=False,
         warn=None, crit=None, query=lambda: _count(db_cosmina, "cantieri")),
    dict(cat="Compliance", el="Garanzie pending", desc="cosmina_warranty_pending da processare",
         fonte="COSMINA", coll="cosmina_warranty_pending", freq="settimanale", sev="warning", badge=False,
         warn=50, crit=200,
         query=lambda: _count(db_cosmina, "cosmina_warranty_pending")),
    dict(cat="Compliance", el="Dossier riqualificazione", desc="dossier_riqualificazione aperti",
         fonte="COSMINA", coll="dossier_riqualificazione", freq="mensile", sev="info", badge=False,
         warn=None, crit=None, query=lambda: _count(db_cosmina, "dossier_riqualificazione")),

    # ─── MPLS / PREVENTIVI ─────────────────────────────────────
    dict(cat="Fatturazione", el="MPLS in bozza", desc="Multi-Preventivi Lavori Straordinari non ancora inviati",
         fonte="Guazzotti", coll="mpls", freq="settimanale", sev="warning", badge=True,
         warn=3, crit=10,
         query=lambda: _count(db_gzt, "mpls", lambda v: str(v.get("status", "")).lower() == "bozza")),

    # ─── NOTIFICHE / PROCESSING ────────────────────────────────
    dict(cat="Infrastruttura", el="Notifiche interne pending", desc="cosmina_notifiche non lette/processate",
         fonte="COSMINA", coll="cosmina_notifiche", freq="oraria", sev="info", badge=False,
         warn=None, crit=None, query=None),
    dict(cat="Infrastruttura", el="Push notifications pending", desc="pendingPushNotifications non inviate",
         fonte="COSMINA", coll="pendingPushNotifications", freq="oraria", sev="warning", badge=False,
         warn=5, crit=30,
         query=lambda: _count(db_cosmina, "pendingPushNotifications")),
    dict(cat="Infrastruttura", el="Queue Cosmina (generica)", desc="cosmina_queue docs — coda processing",
         fonte="COSMINA", coll="cosmina_queue", freq="continua", sev="info", badge=False,
         warn=None, crit=None,
         query=lambda: _count(db_cosmina, "cosmina_queue")),
]

# ─── Helper scoring ────────────────────────────────────────────
def stato_from_val(val, warn, crit):
    """Stato: OK, WARNING, CRITICAL basato su soglie."""
    if val is None or isinstance(val, str):
        return "—"
    if crit is not None and val >= crit:
        return "CRITICAL"
    if warn is not None and val >= warn:
        return "WARNING"
    return "OK"

# ─── Breakdown per tecnico ─────────────────────────────────────
def breakdown_per_tecnico():
    """Aggrega KPI per tecnico da bacheca_cards, rti, rtidf."""
    rows = defaultdict(lambda: {"interventi_aperti": 0, "oggi": 0, "scaduti": 0, "da_programmare": 0,
                                "rti_bozza": 0, "rti_definito": 0, "rtidf_inviato": 0})

    # bacheca_cards INTERVENTI (filtro mirato per ridurre docs)
    for d in db_cosmina.collection("bacheca_cards").where(filter=firestore.FieldFilter("listName", "==", "INTERVENTI")).limit(3000).stream():
        v = d.to_dict() or {}
        if v.get("listName") != "INTERVENTI" or v.get("inBacheca") is False:
            continue
        stato = str(v.get("stato", "")).lower()
        if stato in ("chiuso", "completato"):
            continue
        tec = v.get("techName") or (v.get("techNames") or [None])[0] or "(non assegnato)"
        tec = str(tec).upper().strip()
        r = rows[tec]
        r["interventi_aperti"] += 1
        due = _parse_date(v.get("due"))
        if due:
            if due.date() == dt.datetime.utcnow().date():
                r["oggi"] += 1
            elif due < _now():
                r["scaduti"] += 1
        else:
            r["da_programmare"] += 1

    # RTI Guazzotti
    for d in db_gzt.collection("rti").limit(3000).stream():
        v = d.to_dict() or {}
        tec = str(v.get("tecnico_intervento") or v.get("tecnico") or "(non assegnato)").upper().strip()
        if tec in ("", "NONE"):
            tec = "(non assegnato)"
        st = v.get("stato")
        r = rows[tec]
        if st == "bozza":
            r["rti_bozza"] += 1
        elif st == "definito":
            r["rti_definito"] += 1

    # RTIDF Guazzotti
    for d in db_gzt.collection("rtidf").limit(3000).stream():
        v = d.to_dict() or {}
        tec = str(v.get("tecnico_intervento") or v.get("tecnico") or "(non assegnato)").upper().strip()
        if tec in ("", "NONE"):
            tec = "(non assegnato)"
        if v.get("stato") == "inviato":
            rows[tec]["rtidf_inviato"] += 1

    return rows

# ─── Build Excel ───────────────────────────────────────────────
def build_excel(out_path):
    wb = Workbook()

    # stili
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="0369A1")
    border = Border(left=Side(style="thin", color="E2E8F0"),
                    right=Side(style="thin", color="E2E8F0"),
                    top=Side(style="thin", color="E2E8F0"),
                    bottom=Side(style="thin", color="E2E8F0"))
    fill_warn = PatternFill("solid", fgColor="FEF3C7")
    fill_crit = PatternFill("solid", fgColor="FEE2E2")
    fill_ok   = PatternFill("solid", fgColor="DCFCE7")

    # ── Sheet 1: CATALOGO MONITORAGGIO ──
    ws1 = wb.active
    ws1.title = "CATALOGO MONITORAGGIO"
    headers1 = ["Categoria", "Elemento", "Descrizione", "Fonte Dati",
                "Collection", "Frequenza Suggerita", "Severità", "Badge CHRONOS", "Note"]
    ws1.append(headers1)
    for col in range(1, len(headers1) + 1):
        c = ws1.cell(row=1, column=col)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(vertical="center", wrap_text=True)
        c.border = border

    for item in CATALOG:
        ws1.append([
            item["cat"], item["el"], item["desc"], item["fonte"],
            item["coll"], item["freq"], item["sev"],
            "✅" if item["badge"] else "",
            ""
        ])
    widths1 = [15, 42, 52, 11, 28, 13, 11, 15, 30]
    for i, w in enumerate(widths1, start=1):
        ws1.column_dimensions[get_column_letter(i)].width = w
    for row in ws1.iter_rows(min_row=2, max_row=ws1.max_row):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws1.freeze_panes = "A2"

    # ── Sheet 2: NUMERI ATTUALI ──
    ws2 = wb.create_sheet("NUMERI ATTUALI")
    headers2 = ["Elemento", "Categoria", "Valore Attuale", "Soglia Warning", "Soglia Critical", "Stato", "Note"]
    ws2.append(headers2)
    for col in range(1, len(headers2) + 1):
        c = ws2.cell(row=1, column=col)
        c.font = header_font; c.fill = header_fill
        c.alignment = Alignment(vertical="center"); c.border = border

    for item in CATALOG:
        print(f"  · {item['el']:<55} …", end=" ", flush=True)
        val = None
        if item["query"]:
            try:
                val = item["query"]()
            except Exception as e:
                val = f"ERR: {str(e)[:50]}"
        stato = stato_from_val(val if not isinstance(val, str) else None, item["warn"], item["crit"])
        if val is None:
            stato = "TODO"
        print(f"{val}  [{stato}]")
        ws2.append([
            item["el"], item["cat"],
            val if val is not None else "—",
            item["warn"] if item["warn"] is not None else "—",
            item["crit"] if item["crit"] is not None else "—",
            stato,
            ""
        ])
        # colora riga in base a stato
        r = ws2.max_row
        fill = None
        if stato == "CRITICAL": fill = fill_crit
        elif stato == "WARNING": fill = fill_warn
        elif stato == "OK": fill = fill_ok
        if fill:
            for col in range(1, len(headers2) + 1):
                ws2.cell(row=r, column=col).fill = fill
    widths2 = [50, 15, 18, 16, 16, 12, 30]
    for i, w in enumerate(widths2, start=1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row):
        for cell in row:
            cell.border = border
    ws2.freeze_panes = "A2"

    # ── Sheet 3: PER TECNICO ──
    ws3 = wb.create_sheet("PER TECNICO")
    headers3 = ["Tecnico", "Interventi aperti", "Oggi", "Scaduti", "Da programmare",
                "RTI bozza", "RTI definito (senza RTIDF)", "RTIDF inviati (pronti fattura)"]
    ws3.append(headers3)
    for col in range(1, len(headers3) + 1):
        c = ws3.cell(row=1, column=col)
        c.font = header_font; c.fill = header_fill
        c.alignment = Alignment(vertical="center"); c.border = border

    print("\nBreakdown per tecnico…")
    tec_data = breakdown_per_tecnico()
    # ordina per interventi_aperti desc
    tec_sorted = sorted(tec_data.items(),
                        key=lambda x: x[1]["interventi_aperti"] + x[1]["rti_bozza"] + x[1]["rti_definito"],
                        reverse=True)
    for tec, kp in tec_sorted:
        ws3.append([
            tec, kp["interventi_aperti"], kp["oggi"], kp["scaduti"],
            kp["da_programmare"], kp["rti_bozza"], kp["rti_definito"], kp["rtidf_inviato"],
        ])
    widths3 = [28, 16, 8, 10, 18, 14, 24, 28]
    for i, w in enumerate(widths3, start=1):
        ws3.column_dimensions[get_column_letter(i)].width = w
    for row in ws3.iter_rows(min_row=2, max_row=ws3.max_row):
        for cell in row:
            cell.border = border
            if cell.column > 1:
                cell.alignment = Alignment(horizontal="right")
    ws3.freeze_panes = "A2"

    wb.save(out_path)
    return out_path

if __name__ == "__main__":
    out = os.path.join(os.path.dirname(os.path.abspath(__file__)), "report_monitoraggio.xlsx")
    print(f"→ Building {out}")
    build_excel(out)
    print(f"\n✅ Saved: {out}")
